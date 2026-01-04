"""
Excel Report Generator
Creates detailed Excel and TXT reports from parsing results
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Optional
from dataclasses import dataclass

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.telethon_core import ParsingResult, ParsedUser

logger = logging.getLogger(__name__)


@dataclass
class SmartExportResult:
    """Результат умной выгрузки (4 файла)"""
    admins_txt: Optional[Path] = None
    premium_txt: Optional[Path] = None
    regular_txt: Optional[Path] = None
    full_xlsx: Optional[Path] = None
    
    def all_paths(self) -> List[Path]:
        """Получить список всех созданных файлов"""
        return [p for p in [self.admins_txt, self.premium_txt, self.regular_txt, self.full_xlsx] if p]
    
    def cleanup(self):
        """Удалить все созданные файлы"""
        for path in self.all_paths():
            try:
                if path.exists():
                    path.unlink()
            except:
                pass


class ExcelGenerator:
    """Генератор Excel и TXT отчетов"""

    def __init__(self, output_dir: Path = Path("reports")):
        self.output_dir = output_dir
        self.output_dir.mkdir(exist_ok=True)

    def generate_reports(
        self,
        result: ParsingResult,
        parse_type: str,
        time_filter: Optional[str] = None
    ) -> tuple[Optional[Path], Optional[Path]]:
        """
        Создать оба отчета: Excel и TXT (старый метод для совместимости)
        Returns: (excel_path, txt_path)
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            target_name = self._sanitize_filename(result.target_title or "unknown")

            # Создаем Excel отчет
            excel_path = self.output_dir / f"{target_name}_{timestamp}.xlsx"
            self._create_excel_report(result, excel_path, parse_type, time_filter)

            # Создаем TXT отчет
            txt_path = self.output_dir / f"{target_name}_{timestamp}_usernames.txt"
            self._create_txt_report(result, txt_path)

            logger.info(f"Reports generated: {excel_path}, {txt_path}")
            return excel_path, txt_path

        except Exception as e:
            logger.error(f"Error generating reports: {e}", exc_info=True)
            return None, None
    
    def generate_smart_export(
        self,
        result: ParsingResult,
        parse_type: str,
        time_filter: Optional[str] = None,
        include_bio: bool = False,
        include_gender: bool = False
    ) -> SmartExportResult:
        """
        Умная выгрузка: создаёт 4 файла:
        - admins.txt: Список юзернеймов администраторов
        - premium_users.txt: Список юзернеймов премиум-пользователей
        - regular_users.txt: Обычные пользователи
        - full_report.xlsx: Полная таблица со всеми данными
        """
        export_result = SmartExportResult()
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            target_name = self._sanitize_filename(result.target_title or "unknown")
            
            # 1. admins.txt
            admins_path = self.output_dir / f"{target_name}_{timestamp}_admins.txt"
            self._create_admins_txt(result, admins_path)
            export_result.admins_txt = admins_path
            
            # 2. premium_users.txt
            premium_path = self.output_dir / f"{target_name}_{timestamp}_premium.txt"
            self._create_premium_txt(result, premium_path)
            export_result.premium_txt = premium_path
            
            # 3. regular_users.txt
            regular_path = self.output_dir / f"{target_name}_{timestamp}_regular.txt"
            self._create_regular_txt(result, regular_path)
            export_result.regular_txt = regular_path
            
            # 4. full_report.xlsx
            xlsx_path = self.output_dir / f"{target_name}_{timestamp}_full_report.xlsx"
            self._create_full_xlsx(result, xlsx_path, parse_type, time_filter, include_bio, include_gender)
            export_result.full_xlsx = xlsx_path
            
            logger.info(f"Smart export completed: 4 files generated for {target_name}")
            
        except Exception as e:
            logger.error(f"Error in smart export: {e}", exc_info=True)
        
        return export_result
    
    def _create_admins_txt(self, result: ParsingResult, output_path: Path):
        """Создать файл только с админами"""
        admin_usernames = []
        for admin in result.admins:
            if admin.username:
                admin_usernames.append(f"@{admin.username}")
        
        admin_usernames = sorted(set(admin_usernames))
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(f"# Администраторы - {result.target_title}\n")
            f.write(f"# Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"# Всего: {len(admin_usernames)}\n\n")
            for username in admin_usernames:
                f.write(f"{username}\n")
        
        logger.info(f"admins.txt created: {len(admin_usernames)} admins")
    
    def _create_premium_txt(self, result: ParsingResult, output_path: Path):
        """Создать файл только с премиум-пользователями"""
        premium_usernames = []
        admin_ids = {admin.user_id for admin in result.admins}
        
        for user in result.users:
            if user.is_premium and user.username and user.user_id not in admin_ids:
                premium_usernames.append(f"@{user.username}")
        
        premium_usernames = sorted(set(premium_usernames))
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(f"# Premium пользователи - {result.target_title}\n")
            f.write(f"# Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"# Всего: {len(premium_usernames)}\n\n")
            for username in premium_usernames:
                f.write(f"{username}\n")
        
        logger.info(f"premium_users.txt created: {len(premium_usernames)} premium users")
    
    def _create_regular_txt(self, result: ParsingResult, output_path: Path):
        """Создать файл с обычными пользователями (не админы, не премиум)"""
        regular_usernames = []
        admin_ids = {admin.user_id for admin in result.admins}
        
        for user in result.users:
            if not user.is_premium and not user.is_admin and user.username:
                if user.user_id not in admin_ids:
                    regular_usernames.append(f"@{user.username}")
        
        regular_usernames = sorted(set(regular_usernames))
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(f"# Обычные пользователи - {result.target_title}\n")
            f.write(f"# Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"# Всего: {len(regular_usernames)}\n\n")
            for username in regular_usernames:
                f.write(f"{username}\n")
        
        logger.info(f"regular_users.txt created: {len(regular_usernames)} regular users")
    
    def _get_username_display(self, user: ParsedUser) -> str:
        """Получить отображаемый username или заглушку"""
        if user.username:
            return f"@{user.username}"
        else:
            # Вместо пустоты или "Нет" - понятное значение
            return "Скрыт"

    def _get_username_link(self, user: ParsedUser) -> str:
        """Получить ссылку на пользователя или ID"""
        if user.username:
            return f"https://t.me/{user.username}"
        else:
            return f"tg://user?id={user.user_id}"

    def _create_full_xlsx(
        self,
        result: ParsingResult,
        output_path: Path,
        parse_type: str,
        time_filter: Optional[str],
        include_bio: bool = False,
        include_gender: bool = False
    ):
        """Создать полный Excel отчёт с расширенными колонками"""
        admin_ids = {admin.user_id for admin in result.admins}

        # Определяем статус пользователя
        def get_status(user: ParsedUser) -> str:
            if user.user_id in admin_ids or user.is_admin:
                return "Admin"
            elif user.is_premium:
                return "Premium"
            else:
                return "Member"

        # Подготавливаем данные для всех пользователей
        all_data = []

        # Сначала админы
        for admin in result.admins:
            row = {
                "ID": admin.user_id,
                "Username": self._get_username_display(admin),
                "Link": self._get_username_link(admin),
                "Name": self._get_full_name(admin),
                "Status": "Admin"
            }
            if include_bio:
                row["Bio"] = admin.bio or "—"
            if include_gender:
                row["Gender"] = admin.gender or "неизвестно"
            all_data.append(row)

        # Потом все остальные
        for user in result.users:
            if user.user_id in admin_ids:
                continue  # Уже добавлен как админ

            row = {
                "ID": user.user_id,
                "Username": self._get_username_display(user),
                "Link": self._get_username_link(user),
                "Name": self._get_full_name(user),
                "Status": get_status(user)
            }
            if include_bio:
                row["Bio"] = user.bio or "—"
            if include_gender:
                row["Gender"] = user.gender or "неизвестно"
            all_data.append(row)
        
        # Создаём DataFrame
        df_all = pd.DataFrame(all_data)
        
        # Статистика
        stats_data = {
            "Параметр": [
                "Цель парсинга",
                "Тип парсинга",
                "Временной фильтр",
                "Всего пользователей",
                "Администраторов",
                "Premium пользователей",
                "Обычных пользователей",
                "Время парсинга (сек)",
                "Дата создания"
            ],
            "Значение": [
                result.target_title or "N/A",
                parse_type,
                time_filter or "Без фильтра",
                len(result.users),
                len(result.admins),
                len([u for u in result.users if u.is_premium]),
                len([u for u in result.users if not u.is_premium and not u.is_admin]),
                round(result.parsing_time, 2),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
        }
        df_stats = pd.DataFrame(stats_data)
        
        # Записываем в Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            if not df_all.empty:
                df_all.to_excel(writer, sheet_name='All Users', index=False)
            df_stats.to_excel(writer, sheet_name='Statistics', index=False)
        
        # Применяем стили
        self._apply_excel_styles(output_path)
        
        logger.info(f"full_report.xlsx created: {len(all_data)} total records")

    def _create_excel_report(
        self,
        result: ParsingResult,
        output_path: Path,
        parse_type: str,
        time_filter: Optional[str] = None,
        include_bio: bool = False,
        include_gender: bool = False
    ):
        """Создать Excel отчет с админами в топе"""
        
        # Проверяем есть ли био/гендер в данных
        has_bio = include_bio or any(u.bio for u in result.users + result.admins)
        has_gender = include_gender or any(u.gender for u in result.users + result.admins)

        # Подготовка данных для админов (будут в топе)
        admins_data = []
        for admin in result.admins:
            row = {
                "Type": "👑 ADMIN",
                "Username": self._get_username_display(admin),
                "Link": self._get_username_link(admin),
                "UserID": admin.user_id,
                "FullName": self._get_full_name(admin),
                "LastActivityDate": admin.last_activity.strftime("%Y-%m-%d %H:%M:%S") if admin.last_activity else "N/A",
                "MessageCount": admin.message_count if admin.message_count else 0,
                "Phone": admin.phone if admin.phone else "—"
            }
            if has_bio:
                row["Bio"] = admin.bio or "—"
            if has_gender:
                row["Gender"] = admin.gender or "неизвестно"
            admins_data.append(row)

        # Подготовка данных для обычных пользователей
        users_data = []
        for user in result.users:
            if user.is_admin:  # Пропускаем админов - они уже добавлены
                continue

            row = {
                "Type": "User",
                "Username": self._get_username_display(user),
                "Link": self._get_username_link(user),
                "UserID": user.user_id,
                "FullName": self._get_full_name(user),
                "LastActivityDate": user.last_activity.strftime("%Y-%m-%d %H:%M:%S") if user.last_activity else "N/A",
                "MessageCount": user.message_count,
                "Phone": user.phone if user.phone else "—"
            }
            if has_bio:
                row["Bio"] = user.bio or "—"
            if has_gender:
                row["Gender"] = user.gender or "неизвестно"
            users_data.append(row)
        
        # Объединяем: сначала админы, потом пользователи
        all_users_data = admins_data + users_data

        # Подготовка данных для листа Raw Data
        raw_data = []
        for msg in result.raw_messages[:1000]:  # Ограничиваем 1000 сообщениями
            username = msg.get('username')
            raw_data.append({
                "UserID": msg.get("user_id", ""),
                "Username": f"@{username}" if username else "Скрыт",
                "MessageText": msg.get("text", "")[:500],  # Ограничиваем длину
                "Date": msg.get("date", ""),
                "MessageLink": msg.get("message_link", "")
            })

        # Создаем DataFrame'ы
        df_all = pd.DataFrame(all_users_data)  # Все вместе: админы сверху
        df_admins_only = pd.DataFrame(admins_data)  # Только админы отдельно
        df_raw = pd.DataFrame(raw_data)

        # Записываем в Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Лист All Users (админы в топе, помечены)
            if not df_all.empty:
                df_all.to_excel(writer, sheet_name='All Users', index=False)
            else:
                pd.DataFrame({"Примечание": ["Пользователи не найдены"]}).to_excel(
                    writer, sheet_name='All Users', index=False
                )

            # Лист только с админами
            if not df_admins_only.empty:
                df_admins_only.to_excel(writer, sheet_name='Admins Only', index=False)
            else:
                pd.DataFrame({"Примечание": ["Администраторы не найдены"]}).to_excel(
                    writer, sheet_name='Admins Only', index=False
                )

            # Лист Raw Data
            if not df_raw.empty:
                df_raw.to_excel(writer, sheet_name='Raw Data', index=False)

            # Лист со статистикой
            stats_data = {
                "Параметр": [
                    "Цель парсинга",
                    "Тип парсинга",
                    "Временной фильтр",
                    "Всего пользователей найдено",
                    "Администраторов найдено",
                    "Сообщений просканировано",
                    "Время парсинга (сек)",
                    "Дата создания отчета"
                ],
                "Значение": [
                    result.target_title or "N/A",
                    "Канал (комментарии)" if parse_type == "channel" else "Чат (группа)",
                    time_filter or "Без фильтра",
                    len(result.users),
                    len(result.admins),
                    result.total_messages_scanned,
                    round(result.parsing_time, 2),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ]
            }
            df_stats = pd.DataFrame(stats_data)
            df_stats.to_excel(writer, sheet_name='Statistics', index=False)

        # Применяем стили
        self._apply_excel_styles(output_path)

    def _apply_excel_styles(self, excel_path: Path):
        """Применить стили к Excel файлу"""
        try:
            wb = load_workbook(excel_path)

            # Стили
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Стилизация заголовков
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # Автоширина колонок
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)  # Максимум 50
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Границы для всех ячеек
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = border

            wb.save(excel_path)
            logger.info("Excel styles applied successfully")

        except Exception as e:
            logger.warning(f"Could not apply Excel styles: {e}")

    def _create_txt_report(self, result: ParsingResult, output_path: Path):
        """Создать TXT отчет со списком юзернеймов (админы сверху)"""
        try:
            admin_usernames = []
            user_usernames = []

            # Собираем юзернеймы админов
            for admin in result.admins:
                if admin.username:
                    admin_usernames.append(f"@{admin.username}")
            
            # Собираем юзернеймы обычных пользователей
            for user in result.users:
                if user.username and not user.is_admin:
                    user_usernames.append(f"@{user.username}")

            # Убираем дубликаты
            admin_usernames = sorted(set(admin_usernames))
            user_usernames = sorted(set(user_usernames))
            
            # Исключаем админов из обычных пользователей (если пересекаются)
            user_usernames = [u for u in user_usernames if u not in admin_usernames]

            # Записываем в файл
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(f"# Список юзернеймов - {result.target_title}\n")
                f.write(f"# Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"# Админов: {len(admin_usernames)} | Пользователей: {len(user_usernames)}\n")
                f.write(f"# Всего: {len(admin_usernames) + len(user_usernames)}\n\n")
                
                # Сначала админы
                if admin_usernames:
                    f.write("# ===== АДМИНИСТРАТОРЫ =====\n")
                    for username in admin_usernames:
                        f.write(f"{username}\n")
                    f.write("\n")
                
                # Потом пользователи
                if user_usernames:
                    f.write("# ===== ПОЛЬЗОВАТЕЛИ =====\n")
                    for username in user_usernames:
                        f.write(f"{username}\n")

            logger.info(f"TXT report created: {len(admin_usernames)} admins, {len(user_usernames)} users")

        except Exception as e:
            logger.error(f"Error creating TXT report: {e}")

    def _get_full_name(self, user: ParsedUser) -> str:
        """Получить полное имя пользователя"""
        parts = []
        if user.first_name:
            parts.append(user.first_name)
        if user.last_name:
            parts.append(user.last_name)

        if parts:
            return " ".join(parts)
        elif user.username:
            return f"@{user.username}"
        else:
            return f"ID: {user.user_id}"

    def _sanitize_filename(self, name: str) -> str:
        """Очистить имя файла от недопустимых символов"""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            name = name.replace(char, '_')

        # Ограничиваем длину
        if len(name) > 50:
            name = name[:50]

        return name


# Глобальный экземпляр
excel_generator = ExcelGenerator()
